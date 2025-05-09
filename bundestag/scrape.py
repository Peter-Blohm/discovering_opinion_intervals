from time import sleep

import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
import os
import unicodedata

GERMAN_LETTERS = "äöüÄÖÜß"


def strip_non_german_accents(text: str) -> str:
    if text is None:
        return text

    out = []
    for ch in str(text):
        if ch in GERMAN_LETTERS:
            out.append(ch)
        else:
            # compatibility-decompose the code-point, then drop
            # every combining mark (category "Mn")
            decomp = unicodedata.normalize("NFKD", ch)
            out.append("".join(c for c in decomp if unicodedata.category(c) != "Mn"))
    # NFC re-compose (tidies up e.g. "A\u030A" → "Å")
    return unicodedata.normalize("NFC", "".join(out))


links = []
ljnks = [1]
count = 0
while len(ljnks) > 0:
    INDEX_URL = f"https://www.bundestag.de/ajax/filterlist/de/parlament/plenum/abstimmung/liste/462112-462112?limit=30&noFilterSet=true&offset={30*count}"  # "Namenslisten der Abstimmungen" section
    count = count + 1
    resp = requests.get(INDEX_URL)
    soup = BeautifulSoup(resp.text, "html.parser")
    ljnks = []
    for a in soup.find_all("a", href=True):
        if a["href"].endswith(".xlsx") or a["href"].endswith(".xls"):
            ljnks.append(a["href"])
    print(f"fetched {len(ljnks)} links")
    sleep(1)

    links.extend(ljnks)

os.makedirs("abstimmungen_xlsx", exist_ok=True)

all_votes = []
prayers = 0
for link in links:
    sleep(1)
    print(link, prayers)
    prayers = prayers + 1
    filename = os.path.join("abstimmungen_xlsx", os.path.basename(link))

    if not os.path.exists(filename):
        file_resp = requests.get(link)
        with open(filename, "wb") as f:
            f.write(file_resp.content)

    wb = load_workbook(filename, read_only=True)
    ws = wb.active
    title = ws["A1"].value
    parts = title.split("–")
    date_str = parts[0].replace("Abstimmung am", "").strip()
    subject = parts[-1].replace("Abstimmung über", "").strip()

    df = pd.read_excel(
        filename, header=2, usecols=["Name", "Stimme"], engine="openpyxl"
    )
    df["vote_date"] = pd.to_datetime(date_str, dayfirst=True)
    df["vote_subject"] = subject
    all_votes.append(df)

result_df = pd.concat(all_votes, ignore_index=True)
result_df.to_excel("all_namentliche_abstimmungen.xlsx", index=False)

print(
    f"Downloaded and compiled {len(result_df)} vote records into 'all_namentliche_abstimmungen.xlsx'"
)

import pandas as pd
import glob
import os

folder = "abstimmungen_xlsx"

# pattern to match both .xls and .xlsx
files = glob.glob(os.path.join(folder, "*.xls*"))

all_dfs = []

for filepath in files:
    # read Excel
    df = pd.read_excel(filepath, dtype={"ja": int, "nein": int})

    df = df[(df["ja"] == 1) | (df["nein"] == 1)]

    df["ja/nein"] = df.apply(lambda row: "ja" if row["ja"] == 1 else "nein", axis=1)

    df = df[["Name", "Vorname", "Fraktion/Gruppe", "ja/nein", "Wahlperiode"]].copy()

    df["Name"] = df["Name"].map(lambda x: x.rsplit(" ", 1)[0])
    df["Vorname"] = df["Vorname"].map(lambda x: x.rsplit(" ", 1)[0])
    df["janein"] = df["ja/nein"]
    df["Bezeichnung"] = (df["Name"] + " " + df["Vorname"]).apply(
        strip_non_german_accents
    )

    df.insert(0, "filename", os.path.basename(filepath))

    all_dfs.append(df)

# concatenate all and write out
result = pd.concat(all_dfs, ignore_index=True)
result.to_csv("all_votes.csv", index=False)
